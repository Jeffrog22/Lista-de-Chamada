import pandas as pd
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta
import calendar
from pydantic import BaseModel
from typing import List, Dict, Tuple, Optional
import time, os
import io
from urllib.parse import unquote

# --- INICIALIZAÇÃO DO APP FASTAPI ---
app = FastAPI(
    title="API Gerenciador de Chamadas",
    description="Fornece dados da planilha de alunos e turmas.",
)

# --- CONFIGURAÇÃO DE CORS ---
# Permite que o frontend (rodando em outra porta/endereço) acesse esta API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Em produção, restrinja para o domínio do seu frontend
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- CONSTANTES E FUNÇÕES AUXILIARES ---
NOME_ARQUIVO = 'chamadaBelaVista.xlsx'
TEMPLATE_RELATORIO = 'relatorioChamada.xlsx'
CACHE_EXPIRATION_SECONDS = 60  # Recarrega os dados do Excel a cada 60 segundos
_cache: Dict[str, any] = {"data": None, "timestamp": 0}

def calcular_idade(data_nascimento):
    """Calcula a idade a partir da data de nascimento."""
    if pd.isna(data_nascimento) or not isinstance(data_nascimento, (datetime, pd.Timestamp)):
        return None
    hoje = datetime.now()
    # Calcula a idade subtraindo o ano de nascimento do ano atual.
    # Em seguida, subtrai 1 se o aniversário deste ano ainda não ocorreu.
    idade = hoje.year - data_nascimento.year - ((hoje.month, hoje.day) < (data_nascimento.month, data_nascimento.day))
    return idade

def definir_categoria_por_idade(idade, df_categorias):
    """Define a categoria com base na idade, usando a tabela de categorias fornecida."""
    if pd.isna(idade) or idade is None:
        return "Não definida"
    
    # Itera sobre as regras de categoria carregadas da planilha
    for _, linha in df_categorias.iterrows():
        idade_min = linha.get('Idade Mínima', 0)
        idade_max = linha.get('Idade Máxima', float('inf'))
        if idade_min <= idade <= idade_max:
            return linha.get('Categoria', 'Não definida')
            
    return "Não definida"

def get_dados_cached() -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Carrega os dados da planilha Excel, usando um cache em memória para evitar
    leituras repetidas do arquivo a cada requisição.
    """
    now = time.time()
    # Verifica se o cache expirou ou se o arquivo foi modificado
    file_mod_time = os.path.getmtime(NOME_ARQUIVO) if os.path.exists(NOME_ARQUIVO) else 0
    if now - _cache.get("timestamp", 0) > CACHE_EXPIRATION_SECONDS or file_mod_time > _cache.get("timestamp", 0):
        try:
            xls = pd.ExcelFile(NOME_ARQUIVO, engine='openpyxl')
            df_alunos = pd.read_excel(xls, sheet_name='Alunos').fillna("")
            df_turmas = pd.read_excel(xls, sheet_name='Turmas').fillna("")

            # Carrega categorias ou cria um DF vazio se a aba não existir
            if 'Categorias' in xls.sheet_names:
                df_categorias = pd.read_excel(xls, sheet_name='Categorias')
            else:
                df_categorias = pd.DataFrame(columns=['Categoria', 'Idade Mínima', 'Idade Máxima'])

            # --- CÁLCULO DE IDADE E CATEGORIA ---
            if 'Data de Nascimento' in df_alunos.columns:
                df_alunos['Data de Nascimento'] = pd.to_datetime(df_alunos['Data de Nascimento'], errors='coerce')
                df_alunos['Idade'] = df_alunos['Data de Nascimento'].apply(calcular_idade)
                df_alunos['Categoria'] = df_alunos['Idade'].apply(definir_categoria_por_idade, args=(df_categorias,))
                df_alunos['Idade'] = df_alunos['Idade'].fillna(0).astype(int)
            
            # Carrega registros ou cria um DF vazio se a aba não existir
            if 'Registros' in xls.sheet_names:
                df_registros = pd.read_excel(xls, sheet_name='Registros')
            else:
                df_registros = pd.DataFrame(columns=['Nome'])

            # Carrega justificativas ou cria um DF vazio se a aba não existir
            if 'Justificativas' in xls.sheet_names:
                df_justificativas = pd.read_excel(xls, sheet_name='Justificativas').fillna("")
            else:
                df_justificativas = pd.DataFrame(columns=['Nome', 'Data', 'Motivo'])

            # Carrega exclusões ou cria um DF vazio se a aba não existir
            if 'Exclusões' in xls.sheet_names:
                df_exclusoes = pd.read_excel(xls, sheet_name='Exclusões').fillna("")
            else:
                df_exclusoes = pd.DataFrame(columns=['Nome', 'Turma', 'Horário', 'Professor', 'Data Exclusão'])

            _cache["data"] = (df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes)
            _cache["timestamp"] = now # Usa 'now' para o timestamp do cache
        except FileNotFoundError:
            raise HTTPException(status_code=500, detail=f"Arquivo '{NOME_ARQUIVO}' não encontrado no servidor.")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Ocorreu um erro crítico ao ler a planilha: {e}")

    return _cache["data"]


def formatar_horario(horario):
    """Formata um objeto de tempo, string ou número para o formato 00h00."""
    if pd.isna(horario):
        return ""
    if isinstance(horario, (datetime, pd.Timestamp, pd.Timedelta)):
        return horario.strftime('%Hh%M')
    horario_str = str(horario).split('.')[0]
    try:
        return datetime.strptime(horario_str.zfill(4), '%H%M').strftime('%Hh%M')
    except ValueError:
        return horario_str

# --- MODELOS DE DADOS (PYDANTIC) ---
class ChamadaPayload(BaseModel):
    """Define a estrutura dos dados de chamada que o frontend enviará."""
    registros: Dict[str, Dict[str, str]]  # Ex: {"Nome Aluno": {"dd/mm/yyyy": "c"}}

class JustificativaPayload(BaseModel):
    """Define a estrutura para salvar uma justificativa."""
    Nome: str
    Data: str
    Motivo: str
    
class RelatorioRequest(BaseModel):
    turma: str
    horario: str
    professor: str
    mes: int
    ano: int

class AlunoPayload(BaseModel):
    """Define a estrutura dos dados de um novo aluno que o frontend enviará."""
    Nome: str
    Aniversario: str
    Gênero: str = ""
    Telefone: str = ""
    Turma: str
    Horário: str
    Professor: str
    ParQ: str = ""
    Nível: str = ""
    Categoria: str = ""

class TurmaNivelPayload(BaseModel):
    """Define a estrutura para atualizar o nível de uma turma."""
    turma: str
    horario: str
    professor: str
    novo_nivel: str

class TurmaPayload(BaseModel):
    """Define a estrutura de uma turma para criação/edição."""
    Turma: str
    Horário: str
    Professor: str
    Nível: str = ""
    Atalho: str = ""
    Data_Inicio: str = ""

class TurmaEditPayload(BaseModel):
    """Estrutura para identificar e atualizar uma turma."""
    old_turma: str
    old_horario: str
    old_professor: str
    new_data: TurmaPayload

# --- ENDPOINTS DA API ---

@app.get("/")
def root():
    """Endpoint raiz para verificar se a API está no ar."""
    return {"status": "API do Gerenciador de Chamadas está online"}

@app.get("/api/filtros")
def obter_opcoes_de_filtro():
    """Retorna listas de opções únicas para os filtros do frontend."""
    df_alunos, df_turmas, df_registros, df_categorias, _, _ = get_dados_cached()
    
    turmas = df_turmas['Turma'].unique().tolist()
    
    # Formata os horários para exibição
    df_turmas['Horario_Formatado'] = df_turmas['Horário'].apply(formatar_horario)
    horarios = df_turmas['Horario_Formatado'].unique().tolist()
    
    professores = df_turmas['Professor'].unique().tolist()
    categorias = df_alunos['Categoria'].unique().tolist() if 'Categoria' in df_alunos.columns else []
    niveis = df_alunos['Nível'].unique().tolist() if 'Nível' in df_alunos.columns else []
    
    # Identifica os anos presentes nas colunas de data do Excel (aba Registros)
    anos_disponiveis = {datetime.now().year}
    for col in df_registros.columns:
        # Verifica se a coluna tem formato de data dd/mm/yyyy
        if isinstance(col, str) and len(col) == 10 and col[2] == '/' and col[5] == '/':
            try:
                ano_col = int(col.split('/')[2])
                anos_disponiveis.add(ano_col)
            except ValueError:
                pass
    
    meses_pt = [
        {"valor": 1, "nome": "Janeiro"},
        {"valor": 2, "nome": "Fevereiro"},
        {"valor": 3, "nome": "Março"},
        {"valor": 4, "nome": "Abril"},
        {"valor": 5, "nome": "Maio"},
        {"valor": 6, "nome": "Junho"},
        {"valor": 7, "nome": "Julho"},
        {"valor": 8, "nome": "Agosto"},
        {"valor": 9, "nome": "Setembro"},
        {"valor": 10, "nome": "Outubro"},
        {"valor": 11, "nome": "Novembro"},
        {"valor": 12, "nome": "Dezembro"}
    ]

    return {
        "turmas": turmas,
        "horarios": horarios,
        "professores": professores,
        "categorias": sorted([cat for cat in categorias if cat != "Não definida"]),
        "niveis": sorted([n for n in niveis if n != ""]) if niveis is not None else [],
        "anos": sorted(list(anos_disponiveis), reverse=True),
        "meses": meses_pt
    }

@app.get("/api/all-alunos")
def get_all_alunos():
    """Retorna a lista completa de alunos."""
    df_alunos, _, _, _, _, _ = get_dados_cached()
    # Formata o horário para exibição consistente
    df_alunos['Horário'] = df_alunos['Horário'].apply(formatar_horario)
    return df_alunos.to_dict(orient='records')

@app.get("/api/all-turmas")
def get_all_turmas():
    """Retorna a lista completa de turmas."""
    df_alunos, df_turmas, _, _, _, _ = get_dados_cached()

    # Formata os horários em ambos os dataframes para garantir a correspondência
    df_alunos['Horario_Formatado'] = df_alunos['Horário'].apply(formatar_horario)
    df_turmas['Horario_Formatado'] = df_turmas['Horário'].apply(formatar_horario)

    # Calcula a contagem de alunos por turma/horário/professor
    student_counts = df_alunos.groupby(['Turma', 'Horario_Formatado', 'Professor']).size().reset_index(name='qtd.')

    # Junta a contagem de volta ao dataframe de turmas
    df_turmas_com_qtd = pd.merge(
        df_turmas,
        student_counts,
        on=['Turma', 'Horario_Formatado', 'Professor'],
        how='left'
    )
    
    # Preenche com 0 turmas que não têm alunos cadastrados e converte para inteiro
    df_turmas_com_qtd['qtd.'] = df_turmas_com_qtd['qtd.'].fillna(0).astype(int)

    # Usa o horário formatado para a resposta e remove colunas auxiliares
    df_turmas_com_qtd = df_turmas_com_qtd.rename(columns={"Horario_Formatado": "Horário"}).drop(columns=['Horário_x', 'Horário_y'], errors='ignore')
    
    # Reordena as colunas: move 'qtd.', 'Atalho' e 'Data de Início' para o final
    cols = df_turmas_com_qtd.columns.tolist()
    cols_final = ['qtd.', 'Atalho']
    
    # Reconstrói a lista: colunas normais + colunas finais (se existirem)
    new_order = [c for c in cols if c not in cols_final] + [c for c in cols_final if c in cols]
    df_turmas_com_qtd = df_turmas_com_qtd[new_order]

    return df_turmas_com_qtd.to_dict(orient='records')

@app.get("/api/categorias")
def get_all_categorias():
    """Retorna a lista completa de categorias com suas regras de idade."""
    _, _, _, df_categorias, _, _ = get_dados_cached()
    return df_categorias.to_dict(orient='records')


@app.get("/api/alunos")
def obter_alunos_filtrados(
    turma: str = Query(...),
    horario: str = Query(...),
    professor: str = Query(...),
    mes: int = Query(...),
    ano: Optional[int] = Query(None)
):
    """
    Retorna a lista de alunos e os registros de presença para um determinado mês e ano.
    """
    df_alunos, _, df_registros, _, df_justificativas, _ = get_dados_cached()
    ano_vigente = ano if ano else datetime.now().year

    # --- Lógica para gerar as datas de aula (adaptada do Streamlit) ---
    dias_da_semana_validos = set()
    nome_turma_lower = turma.lower()
    
    dias_map = {
        "segunda": 0, "terça": 1, "terca": 1, 
        "quarta": 2, "quinta": 3, "sexta": 4, 
        "sábado": 5, "sabado": 5, "domingo": 6
    }
    
    for dia_nome, dia_idx in dias_map.items():
        if dia_nome in nome_turma_lower:
            dias_da_semana_validos.add(dia_idx)
            
    if not dias_da_semana_validos:
        dias_da_semana_validos = list(range(7)) # Fallback: todos os dias se não identificar
    else:
        dias_da_semana_validos = sorted(list(dias_da_semana_validos))

    try:
        dias_no_mes = calendar.monthrange(ano_vigente, mes)[1]
        datas_mes_todas = [datetime(ano_vigente, mes, dia) for dia in range(1, dias_no_mes + 1)]
        datas_mes_filtradas = [data for data in datas_mes_todas if data.weekday() in dias_da_semana_validos]
        datas_mes_str = [data.strftime('%d/%m/%Y') for data in datas_mes_filtradas]
    except ValueError:
        raise HTTPException(status_code=400, detail="Mês inválido.")

    # Garante que as colunas de data existam no df_registros
    for data_str in datas_mes_str:
        if data_str not in df_registros.columns:
            df_registros[data_str] = ""

    # Preenche valores nulos com string vazia para evitar problemas com JSON
    df_registros = df_registros.fillna("")

    # Para garantir a correspondência, criamos uma coluna de horário formatado
    df_alunos['Horario_Formatado'] = df_alunos['Horário'].apply(formatar_horario)

    # Aplica os filtros
    alunos_filtrados = df_alunos[
        (df_alunos['Turma'] == turma) &
        (df_alunos['Horario_Formatado'] == horario) &
        (df_alunos['Professor'] == professor)
    ]

    # Junta os alunos filtrados com seus registros de presença
    alunos_com_registros = pd.merge(
        alunos_filtrados[['Turma', 'Horario_Formatado', 'Professor', 'Nível', 'Nome', 'Idade', 'Categoria', 'Whatsapp', 'ParQ', 'Data de Nascimento']],
        df_registros[['Nome'] + datas_mes_str],
        on='Nome',
        how='left'
    ).fillna("")

    # Renomeia a coluna de horário para o frontend
    alunos_com_registros = alunos_com_registros.rename(columns={"Horario_Formatado": "Horário"})
    
    # Renomeia Data de Nascimento para Aniversario para manter consistência com o frontend/relatório
    if 'Data de Nascimento' in alunos_com_registros.columns:
        alunos_com_registros = alunos_com_registros.rename(columns={'Data de Nascimento': 'Aniversario'})

    # --- PROCESSAMENTO DE JUSTIFICATIVAS ---
    # Filtra as justificativas para o mês e ano solicitados e anexa ao aluno
    if not df_justificativas.empty:
        df_just = df_justificativas.copy()
        # Garante que a coluna Data seja datetime
        df_just['Data_dt'] = pd.to_datetime(df_just['Data'], dayfirst=True, errors='coerce')
        
        # Filtra pelo mês/ano
        mask = (df_just['Data_dt'].dt.month == mes) & (df_just['Data_dt'].dt.year == ano_vigente)
        df_just_mes = df_just[mask].copy()
        
        if not df_just_mes.empty:
            # Ordena por data para garantir cronologia no histórico
            df_just_mes = df_just_mes.sort_values('Data_dt')

            # Cria string formatada "DD - Motivo"
            df_just_mes['Dia'] = df_just_mes['Data_dt'].dt.day.apply(lambda x: f"{x:02d}")
            df_just_mes['Texto'] = df_just_mes['Dia'] + " - " + df_just_mes['Motivo'].astype(str)
            
            # Agrupa por nome e junta as justificativas com quebra de linha
            justificativas_agg = df_just_mes.groupby('Nome')['Texto'].apply(lambda x: '\n'.join(x)).reset_index()
            justificativas_agg.rename(columns={'Texto': 'Justificativas'}, inplace=True)
            
            # Merge com o resultado final
            alunos_com_registros = pd.merge(alunos_com_registros, justificativas_agg, on='Nome', how='left')
            alunos_com_registros['Justificativas'] = alunos_com_registros['Justificativas'].fillna("")
        else:
            alunos_com_registros['Justificativas'] = ""
    else:
        alunos_com_registros['Justificativas'] = ""

    return {
        "datas": datas_mes_str,
        "alunos": alunos_com_registros.to_dict(orient='records')
    }

@app.get("/api/relatorio/frequencia")
def obter_relatorio_frequencia(dias: int = 30):
    """Calcula e retorna as métricas de frequência para um período em dias."""
    try:
        _, _, df_registros, _, _, _ = get_dados_cached()
    except Exception:
        return {"error": "Nenhum registro encontrado."}

    if df_registros.empty:
        return {"error": "Nenhum registro encontrado."}

    hoje = datetime.now()
    data_inicio = hoje - timedelta(days=dias) # Corrigido de datetime para timedelta
    colunas_identificacao = ['Nome', 'Turma', 'Horário', 'Professor', 'Nível']
    colunas_datas = [col for col in df_registros.columns if col not in colunas_identificacao]

    datas_relevantes = []
    for data_str in colunas_datas:
        try:
            data_col = datetime.strptime(data_str, '%d/%m/%Y')
            if data_inicio.date() <= data_col.date() <= hoje.date():
                datas_relevantes.append(data_str)
        except (ValueError, TypeError):
            continue

    if not datas_relevantes:
        return {"error": f"Nenhum registro de chamada nos últimos {dias} dias."}

    df_relatorio = df_registros[['Nome'] + datas_relevantes].copy().fillna('')
    df_relatorio.set_index('Nome', inplace=True)

    total_aulas = len(datas_relevantes)
    presencas = (df_relatorio == 'c').sum(axis=1)
    faltas = (df_relatorio == 'f').sum(axis=1)
    justificadas = (df_relatorio == 'j').sum(axis=1)
    aulas_consideradas = presencas + faltas
    frequencia_percentual = (presencas / aulas_consideradas.replace(0, 1)) * 100

    df_resultado = pd.DataFrame({
        'Total de Aulas no Período': total_aulas, 'Presenças (C)': presencas,
        'Faltas (F)': faltas, 'Faltas Justificadas (J)': justificadas,
        'Frequência (%)': frequencia_percentual.round(2)
    })
    return df_resultado.reset_index().to_dict(orient='records')

@app.get("/api/relatorio/excel")
def gerar_relatorio_excel_endpoint(
    turma: str = Query(...),
    horario: str = Query(...),
    professor: str = Query(...),
    mes: int = Query(...),
    ano: int = Query(...)
):
    """
    Gera um arquivo Excel baseado no template 'relatorioChamada.xlsx' preenchido com os dados da turma.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="A biblioteca 'openpyxl' é necessária no backend.")

    if not os.path.exists(TEMPLATE_RELATORIO):
        raise HTTPException(status_code=404, detail=f"Template '{TEMPLATE_RELATORIO}' não encontrado no servidor.")

    # 1. Obter dados
    dados_api = obter_alunos_filtrados(turma, horario, professor, mes, ano)
    alunos = dados_api.get('alunos', [])
    datas_str = dados_api.get('datas', [])
    
    # 2. Carregar Template
    try:
        wb = load_workbook(TEMPLATE_RELATORIO)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler o template: {e}")

    # 3. Preencher Cabeçalho (Aba Turmas / Interface)
    # B3: Professor, B4: Turma, B5: Horário
    ws['B3'] = professor
    ws['B4'] = turma
    ws['B5'] = horario
    
    # E5: Mês selecionado (Ex: 10/2025)
    ws['E5'] = f"{mes:02d}/{ano}"

    # E6: Datas do mês (Cabeçalho das colunas de presença)
    # Começa na coluna 5 (E)
    col_inicio_datas = 5
    for i, data in enumerate(datas_str):
        # data vem como "dd/mm/yyyy", pegamos apenas o dia "dd" ou a data curta
        dia = data.split('/')[0]
        cell = ws.cell(row=6, column=col_inicio_datas + i)
        cell.value = dia
        cell.alignment = Alignment(horizontal='center')

    # 4. Preencher Linhas de Alunos (A partir da linha 7)
    linha_inicial = 7
    for idx, aluno in enumerate(alunos):
        linha_atual = linha_inicial + idx
        
        # A7: Nome
        ws.cell(row=linha_atual, column=1, value=aluno.get('Nome', ''))
        
        # B7: Whatsapp
        ws.cell(row=linha_atual, column=2, value=aluno.get('Whatsapp', ''))
        
        # C7: ParQ
        ws.cell(row=linha_atual, column=3, value=aluno.get('ParQ', ''))
        
        # D7: Aniversário (Formatado)
        aniversario = aluno.get('Aniversario', '')
        if isinstance(aniversario, (datetime, pd.Timestamp)):
            aniversario = aniversario.strftime('%d/%m/%Y')
        elif aniversario and isinstance(aniversario, str):
             aniversario = aniversario.split('T')[0].split(' ')[0]
             try:
                 dt_aniv = datetime.strptime(aniversario, '%Y-%m-%d')
                 aniversario = dt_aniv.strftime('%d/%m/%Y')
             except ValueError:
                 pass
        ws.cell(row=linha_atual, column=4, value=aniversario)

        # E7 em diante: Registros de Presença
        for i, data in enumerate(datas_str):
            status = aluno.get(data, "")
            # Mapeia códigos para visualização se necessário, ou mantém c/f/j
            ws.cell(row=linha_atual, column=col_inicio_datas + i, value=status).alignment = Alignment(horizontal='center')

        # M7: Nível (Coluna 13)
        ws.cell(row=linha_atual, column=13, value=aluno.get('Nível', ''))

    # 5. Salvar em memória e retornar stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Relatorio_{turma}_{mes}_{ano}.xlsx".replace(" ", "_").replace("/", "-")
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/api/relatorio/excel_consolidado")
def gerar_relatorio_excel_consolidado(requests_list: List[RelatorioRequest]):
    """
    Gera um único arquivo Excel com múltiplas abas (uma para cada turma solicitada),
    baseado no template 'relatorioChamada.xlsx'.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="A biblioteca 'openpyxl' é necessária no backend.")

    if not os.path.exists(TEMPLATE_RELATORIO):
        raise HTTPException(status_code=404, detail=f"Template '{TEMPLATE_RELATORIO}' não encontrado no servidor.")

    # 1. Carregar Template
    try:
        wb = load_workbook(TEMPLATE_RELATORIO)
        template_sheet = wb.active
        template_sheet.title = "Template"
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler o template: {e}")

    # 2. Processar cada solicitação
    for req in requests_list:
        # Obter dados
        dados_api = obter_alunos_filtrados(req.turma, req.horario, req.professor, req.mes, req.ano)
        alunos = dados_api.get('alunos', [])
        datas_str = dados_api.get('datas', [])

        # Criar nova aba copiando o template
        # Limpa caracteres inválidos para nome de aba Excel
        safe_turma = "".join([c for c in req.turma if c.isalnum() or c in (' ', '-', '_')])[:20]
        safe_horario = req.horario.replace(':', 'h')
        sheet_title = f"{safe_turma} {safe_horario}"[:30]
        
        # Se a aba já existir, adiciona sufixo para evitar erro
        count = 1
        original_title = sheet_title
        while sheet_title in wb.sheetnames:
            sheet_title = f"{original_title}_{count}"
            count += 1
        
        ws = wb.copy_worksheet(template_sheet)
        ws.title = sheet_title

        # Preencher Cabeçalho
        ws['B3'] = req.professor
        ws['B4'] = req.turma
        ws['B5'] = req.horario
        ws['E5'] = f"{req.mes:02d}/{req.ano}"

        # Limpa cabeçalhos antigos da linha 6 (das datas até o fim provável) para evitar duplicação
        for c in range(5, 35): 
            ws.cell(row=6, column=c, value="")

        # Preencher Datas (Cabeçalho das colunas de presença)
        col_inicio_datas = 5 # Coluna E
        for i, data in enumerate(datas_str):
            dia = data.split('/')[0]
            cell = ws.cell(row=6, column=col_inicio_datas + i)
            cell.value = dia
            cell.alignment = Alignment(horizontal='center')

        # Definir posição da coluna Nível (Dinâmica: logo após a última data)
        col_nivel = col_inicio_datas + len(datas_str)
        ws.cell(row=6, column=col_nivel, value="Nível").alignment = Alignment(horizontal='center', vertical='center')
        
        # Remover colunas excedentes do template (após a coluna Nível)
        max_col_template = ws.max_column
        if max_col_template > col_nivel:
            ws.delete_cols(col_nivel + 1, max_col_template - col_nivel)
        
        # Rastreia largura máxima para auto-ajuste (inicia com tamanho do cabeçalho)
        max_width_nivel = len("Nível")

        # Preencher Linhas de Alunos
        linha_inicial = 7
        for idx, aluno in enumerate(alunos):
            linha_atual = linha_inicial + idx
            
            ws.cell(row=linha_atual, column=1, value=aluno.get('Nome', ''))
            ws.cell(row=linha_atual, column=2, value=aluno.get('Whatsapp', ''))
            ws.cell(row=linha_atual, column=3, value=aluno.get('ParQ', ''))
            
            # Formata Aniversário
            aniversario = aluno.get('Aniversario', '')
            if isinstance(aniversario, (datetime, pd.Timestamp)):
                aniversario = aniversario.strftime('%d/%m/%Y')
            elif aniversario and isinstance(aniversario, str):
                 aniversario = aniversario.split('T')[0].split(' ')[0]
                 try:
                     dt_aniv = datetime.strptime(aniversario, '%Y-%m-%d')
                     aniversario = dt_aniv.strftime('%d/%m/%Y')
                 except ValueError:
                     pass
            ws.cell(row=linha_atual, column=4, value=aniversario)

            # Registros de Presença
            for i, data in enumerate(datas_str):
                status = aluno.get(data, "")
                ws.cell(row=linha_atual, column=col_inicio_datas + i, value=status).alignment = Alignment(horizontal='center')

            # Nível na coluna dinâmica
            nivel_val = aluno.get('Nível', '')
            ws.cell(row=linha_atual, column=col_nivel, value=nivel_val)
            if nivel_val:
                max_width_nivel = max(max_width_nivel, len(str(nivel_val)))
        
        # Aplica o ajuste de largura na coluna Nível (+ margem visual)
        ws.column_dimensions[get_column_letter(col_nivel)].width = max_width_nivel + 3

    # 3. Remover a aba de template original antes de salvar
    if "Template" in wb.sheetnames:
        wb.remove(wb["Template"])

    # 4. Salvar em memória e retornar stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Relatorio_Consolidado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/api/chamada")
def salvar_chamada(payload: dict):
    """Recebe e salva os registros de chamada na planilha.

    Aceita dois formatos de payload:
    - Estrutura antiga/ideal: {"registros": {"Nome": {"dd/mm/yyyy": "c"}}}
    - Estrutura em lista: {"registros": [{"Nome": "x", "Data": "dd/mm/yyyy", "Status": "c"}, ...]}
    """
    try:
        # Carrega todos os dados do cache para uma reescrita segura
        df_alunos, df_turmas, df_registros_orig, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()
        df_registros = df_registros_orig.copy()

        registros = payload.get("registros")
        if registros is None:
            raise HTTPException(status_code=400, detail="Payload inválido: campo 'registros' ausente.")

        # Converte payload em formato de lista para o formato dict esperado
        if isinstance(registros, list):
            registros_dict: Dict[str, Dict[str, str]] = {}
            for rec in registros:
                nome = rec.get('Nome') or rec.get('name')
                data = rec.get('Data') or rec.get('data')
                status = rec.get('Status') or rec.get('status')
                if not nome or not data:
                    continue
                registros_dict.setdefault(nome, {})[data] = status
            registros = registros_dict

        if not isinstance(registros, dict):
            raise HTTPException(status_code=400, detail="Formato de 'registros' inválido.")

        for nome_aluno, registros_data in registros.items():
            if nome_aluno not in df_registros['Nome'].values:
                nova_linha = pd.DataFrame([{'Nome': nome_aluno}])
                df_registros = pd.concat([df_registros, nova_linha], ignore_index=True)

            # Garante que o índice seja encontrado após uma possível concatenação
            idx_registro = df_registros[df_registros['Nome'] == nome_aluno].index

            for data, status in registros_data.items():
                if data not in df_registros.columns:
                    df_registros[data] = pd.NA
                df_registros.loc[idx_registro, data] = status if status else pd.NA

        # Reescreve o arquivo Excel inteiro com os dados atualizados.
        try:
            with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer: # type: ignore
                df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
                df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
                df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
                df_registros.to_excel(writer, sheet_name='Registros', index=False)
                df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
                df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)
        except PermissionError:
            raise HTTPException(status_code=500, detail=f"Erro de permissão. O arquivo '{NOME_ARQUIVO}' pode estar aberto em outro programa.")

        # Força a limpeza do cache para que a próxima leitura obtenha os dados salvos
        _cache["timestamp"] = 0

        return {"status": "Chamada salva com sucesso!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar a chamada: {e}")

@app.post("/api/justificativa")
def salvar_justificativa(payload: JustificativaPayload):
    """Salva uma nova justificativa na aba 'Justificativas'."""
    try:
        df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()
        
        # Adiciona a nova justificativa
        nova_linha = pd.DataFrame([payload.dict()])
        df_justificativas = pd.concat([df_justificativas, nova_linha], ignore_index=True)
        
        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)
            
        _cache["timestamp"] = 0
        return {"status": "Justificativa salva com sucesso"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar justificativa: {e}")


# --- NOVO ENDPOINT PARA ADICIONAR ALUNO ---
@app.post("/api/aluno")
def adicionar_aluno(aluno_data: AlunoPayload):
    """Adiciona um novo aluno à planilha 'Alunos'."""
    try:
        # Carrega os dados atuais para garantir que não estamos sobrescrevendo nada
        df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()

        # Verifica se o aluno já existe (pelo nome)
        if aluno_data.Nome in df_alunos['Nome'].values:
            raise HTTPException(
                status_code=409, # 409 Conflict
                detail=f"Já existe um aluno com o nome '{aluno_data.Nome}'. Por favor, use um nome diferente."
            )

        # Converte o Pydantic model para um dicionário e depois para um DataFrame
        # Renomeia 'Aniversario' para 'Data de Nascimento' para corresponder à coluna do Excel
        novo_aluno_dict = aluno_data.dict()
        novo_aluno_dict['Data de Nascimento'] = novo_aluno_dict.pop('Aniversario')
        
        # Garante que o nome da coluna de telefone corresponda ao que está no Excel
        if 'Telefone' in novo_aluno_dict:
             novo_aluno_dict['Whatsapp'] = novo_aluno_dict.pop('Telefone')

        novo_aluno_df = pd.DataFrame([novo_aluno_dict])

        # Adiciona a nova linha ao DataFrame de alunos
        df_alunos_atualizado = pd.concat([df_alunos, novo_aluno_df], ignore_index=True)

        # Reescreve o arquivo Excel com a lista de alunos atualizada
        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos_atualizado.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)

        # Força a limpeza do cache para que a próxima leitura inclua o novo aluno
        _cache["timestamp"] = 0

        return {"status": "Aluno adicionado com sucesso!", "aluno": aluno_data.dict()}

    except HTTPException:
        raise # Re-levanta exceções HTTP (como o 409) para que o FastAPI as manipule
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro interno ao adicionar aluno: {e}")

# --- NOVO ENDPOINT PARA ATUALIZAR ALUNO ---
@app.put("/api/aluno/{nome_original}")
def atualizar_aluno(nome_original: str, aluno_data: AlunoPayload):
    """Atualiza os dados de um aluno existente."""
    try:
        # Decodifica o nome da URL (ex: Jos%C3%A9 -> José)
        nome_real = unquote(nome_original)

        # Obtém dados do cache
        df_alunos_cache, df_turmas, df_registros_cache, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()
        
        # Trabalha com cópias para não afetar o cache antes de salvar com sucesso
        df_alunos = df_alunos_cache.copy()
        df_registros = df_registros_cache.copy()

        # Verifica se o aluno existe
        if nome_real not in df_alunos['Nome'].values:
            raise HTTPException(status_code=404, detail=f"Aluno '{nome_real}' não encontrado.")

        # Verifica conflito de nome (se o nome foi alterado e o novo já existe)
        if aluno_data.Nome != nome_real and aluno_data.Nome in df_alunos['Nome'].values:
            raise HTTPException(status_code=409, detail=f"O nome '{aluno_data.Nome}' já está em uso por outro aluno.")

        # Prepara os dados (mapeia campos do payload para colunas do Excel)
        dados_atualizados = aluno_data.dict()
        dados_atualizados['Data de Nascimento'] = dados_atualizados.pop('Aniversario')
        if 'Telefone' in dados_atualizados:
            dados_atualizados['Whatsapp'] = dados_atualizados.pop('Telefone')

        # Atualiza os dados no DataFrame de Alunos
        idx = df_alunos[df_alunos['Nome'] == nome_real].index[0]
        for col, valor in dados_atualizados.items():
            df_alunos.loc[idx, col] = valor

        # Se o nome mudou, atualiza também na planilha de Registros para não perder o histórico
        if aluno_data.Nome != nome_real:
            if 'Nome' in df_registros.columns:
                df_registros.loc[df_registros['Nome'] == nome_real, 'Nome'] = aluno_data.Nome

        # Salva todas as alterações no arquivo Excel
        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)

        # Invalida o cache para forçar recarregamento
        _cache["timestamp"] = 0
        
        return {"status": "Aluno atualizado com sucesso!", "aluno": aluno_data.dict()}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro interno ao atualizar aluno: {e}")

# --- NOVO ENDPOINT PARA EXCLUIR ALUNO (MOVER PARA EXCLUSÕES) ---
@app.delete("/api/aluno/{nome_original}")
def excluir_aluno(nome_original: str):
    """Remove o aluno da lista ativa e o move para a aba 'Exclusões'."""
    try:
        nome_real = unquote(nome_original)
        df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()

        if nome_real not in df_alunos['Nome'].values:
            raise HTTPException(status_code=404, detail=f"Aluno '{nome_real}' não encontrado.")

        # Extrai a linha do aluno
        aluno_row = df_alunos[df_alunos['Nome'] == nome_real].iloc[0].to_dict()
        
        # Adiciona a data de exclusão
        aluno_row['Data Exclusão'] = datetime.now()
        
        # Adiciona à tabela de exclusões
        df_exclusoes = pd.concat([df_exclusoes, pd.DataFrame([aluno_row])], ignore_index=True)
        
        # Remove da tabela de alunos
        df_alunos = df_alunos[df_alunos['Nome'] != nome_real]

        # Salva tudo
        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)

        _cache["timestamp"] = 0
        return {"status": f"Aluno '{nome_real}' movido para Exclusões."}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao excluir aluno: {e}")

# --- NOVO ENDPOINT PARA LISTAR EXCLUSÕES ---
@app.get("/api/exclusoes")
def get_exclusoes():
    """Retorna a lista de alunos excluídos."""
    _, _, _, _, _, df_exclusoes = get_dados_cached()
    # Formata datas se necessário, ou retorna como está
    return df_exclusoes.to_dict(orient='records')

# --- NOVO ENDPOINT PARA RESTAURAR ALUNO ---
@app.post("/api/restaurar")
def restaurar_aluno(aluno_data: AlunoPayload):
    """Restaura um aluno da lista de exclusões para a lista ativa."""
    try:
        df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()
        
        nome_aluno = aluno_data.Nome
        
        # Verifica se já existe na lista ativa (evitar duplicatas)
        if nome_aluno in df_alunos['Nome'].values:
             # Se já existe, apenas removemos da exclusão (assumindo que foi recriado manualmente ou restaurado antes)
             pass
        else:
            # Prepara dados para reinserção
            novo_aluno_dict = aluno_data.dict()
            novo_aluno_dict['Data de Nascimento'] = novo_aluno_dict.pop('Aniversario')
            if 'Telefone' in novo_aluno_dict:
                novo_aluno_dict['Whatsapp'] = novo_aluno_dict.pop('Telefone')
            
            # Adiciona de volta aos alunos
            df_alunos = pd.concat([df_alunos, pd.DataFrame([novo_aluno_dict])], ignore_index=True)

        # Remove da lista de exclusões (remove todas as ocorrências desse nome)
        if 'Nome' in df_exclusoes.columns:
            df_exclusoes = df_exclusoes[df_exclusoes['Nome'] != nome_aluno]

        # Salva tudo
        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)

        _cache["timestamp"] = 0
        return {"status": f"Aluno '{nome_aluno}' restaurado com sucesso."}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao restaurar aluno: {e}")


# --- NOVO ENDPOINT PARA EXCLUIR TURMA ---
@app.delete("/api/turma")
def excluir_turma(
    turma: str = Query(...),
    horario: str = Query(...),
    professor: str = Query(...)
):
    """Exclui uma turma da planilha 'Turmas'."""
    try:
        df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()
        
        # Cria uma cópia para manipulação e formata o horário para localizar a linha correta
        df_turmas_temp = df_turmas.copy()
        df_turmas_temp['Horario_Formatado'] = df_turmas_temp['Horário'].apply(formatar_horario)
        
        # Localiza a turma pelos critérios (Turma, Horário Formatado, Professor)
        mask = (
            (df_turmas_temp['Turma'] == turma) & 
            (df_turmas_temp['Horario_Formatado'] == horario) & 
            (df_turmas_temp['Professor'] == professor)
        )
        
        if not mask.any():
            raise HTTPException(status_code=404, detail="Turma não encontrada para exclusão.")
            
        # Remove as linhas encontradas do DataFrame original
        indices_to_drop = df_turmas_temp[mask].index
        df_turmas = df_turmas.drop(indices_to_drop)
        
        # Salva as alterações no Excel
        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)
            
        # Invalida o cache
        _cache["timestamp"] = 0
        
        return {"status": "Turma excluída com sucesso"}
        
    except HTTPException:
        raise
    except PermissionError:
        raise HTTPException(status_code=500, detail=f"Erro de permissão. O arquivo '{NOME_ARQUIVO}' pode estar aberto.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao excluir turma: {e}")

# --- NOVO ENDPOINT PARA ATUALIZAR NÍVEL DA TURMA ---
@app.put("/api/turma/nivel")
def atualizar_nivel_turma(payload: TurmaNivelPayload):
    """Atualiza o nível de uma turma existente."""
    try:
        df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()
        
        # Cria cópia para manipulação segura e busca
        df_turmas_temp = df_turmas.copy()
        df_turmas_temp['Horario_Formatado'] = df_turmas_temp['Horário'].apply(formatar_horario)
        
        # Localiza a turma
        mask = (
            (df_turmas_temp['Turma'] == payload.turma) & 
            (df_turmas_temp['Horario_Formatado'] == payload.horario) & 
            (df_turmas_temp['Professor'] == payload.professor)
        )
        
        indices = df_turmas_temp[mask].index
        if indices.empty:
            raise HTTPException(status_code=404, detail="Turma não encontrada para atualização.")
            
        # Atualiza o nível no DataFrame original (usando os índices encontrados)
        # Usamos uma cópia para garantir que a escrita no Excel seja limpa
        df_turmas_to_save = df_turmas.copy()
        df_turmas_to_save.loc[indices, 'Nível'] = payload.novo_nivel
        
        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas_to_save.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)
            
        _cache["timestamp"] = 0 # Invalida cache
        return {"status": "Nível atualizado com sucesso"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar nível: {e}")

@app.post("/api/turma")
def adicionar_turma(turma_data: TurmaPayload):
    """Adiciona uma nova turma à planilha 'Turmas'."""
    try:
        df_alunos, df_turmas, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()

        # Verifica duplicidade
        df_check = df_turmas.copy()
        df_check['Horario_Formatado'] = df_check['Horário'].apply(formatar_horario)
        horario_input = formatar_horario(turma_data.Horário)

        exists = ((df_check['Turma'] == turma_data.Turma) & 
                  (df_check['Horario_Formatado'] == horario_input) & 
                  (df_check['Professor'] == turma_data.Professor)).any()

        if exists:
             raise HTTPException(status_code=409, detail="Esta turma já existe.")

        nova_turma = {
            "Turma": turma_data.Turma,
            "Horário": turma_data.Horário,
            "Professor": turma_data.Professor,
            "Nível": turma_data.Nível,
            "Atalho": turma_data.Atalho,
            "Data de Início": turma_data.Data_Inicio
        }
        
        df_turmas = pd.concat([df_turmas, pd.DataFrame([nova_turma])], ignore_index=True)

        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)

        _cache["timestamp"] = 0
        return {"status": "Turma adicionada com sucesso!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao adicionar turma: {e}")

@app.put("/api/turma")
def editar_turma(payload: TurmaEditPayload):
    """Edita uma turma existente."""
    try:
        df_alunos, df_turmas_cache, df_registros, df_categorias, df_justificativas, df_exclusoes = get_dados_cached()
        df_turmas = df_turmas_cache.copy()
        
        df_turmas_temp = df_turmas.copy()
        df_turmas_temp['Horario_Formatado'] = df_turmas_temp['Horário'].apply(formatar_horario)
        
        mask = (
            (df_turmas_temp['Turma'] == payload.old_turma) & 
            (df_turmas_temp['Horario_Formatado'] == payload.old_horario) & 
            (df_turmas_temp['Professor'] == payload.old_professor)
        )
        
        if not mask.any():
            raise HTTPException(status_code=404, detail="Turma original não encontrada.")
            
        idx = df_turmas_temp[mask].index[0]
        
        df_turmas.at[idx, 'Turma'] = payload.new_data.Turma
        df_turmas.at[idx, 'Horário'] = payload.new_data.Horário
        df_turmas.at[idx, 'Professor'] = payload.new_data.Professor
        df_turmas.at[idx, 'Nível'] = payload.new_data.Nível
        df_turmas.at[idx, 'Atalho'] = payload.new_data.Atalho
        df_turmas.at[idx, 'Data de Início'] = payload.new_data.Data_Inicio

        with pd.ExcelWriter(NOME_ARQUIVO, engine='openpyxl') as writer:
            df_alunos.to_excel(writer, sheet_name='Alunos', index=False)
            df_turmas.to_excel(writer, sheet_name='Turmas', index=False)
            df_categorias.to_excel(writer, sheet_name='Categorias', index=False)
            df_registros.to_excel(writer, sheet_name='Registros', index=False)
            df_justificativas.to_excel(writer, sheet_name='Justificativas', index=False)
            df_exclusoes.to_excel(writer, sheet_name='Exclusões', index=False)
            
        _cache["timestamp"] = 0
        return {"status": "Turma atualizada com sucesso!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao editar turma: {e}")

# Para rodar este servidor, use o comando no terminal:
# uvicorn backend:app --reload